"""v2 verifier for browser-web-research-electric-cargo-tricycle-market-report.

Mixed deliverable: xlsx (6 sheets) + md notes + JSON manifest + CSV + decision JSON + 4 screenshots.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from io import StringIO
from pathlib import Path
from typing import Any


XLSX_NAME = "Electric_Cargo_Tricycle_Market_Report.xlsx"
NOTES_NAME = "market_report_notes.md"
MANIFEST_NAME = "evidence_manifest.json"
CSV_NAME = "normalized_comparison.csv"
DECISION_NAME = "channel_decision_matrix.json"
SCREENSHOTS = {"alibaba.png", "amazon.png", "walmart.png", "brand_site.png"}

REQUIRED_SHEETS = ["Alibaba.com", "Amazon", "Walmart", "Brand Sites", "Normalized Comparison", "Opportunity Matrix"]
SHEET_MIN_ROWS = {
    "Alibaba.com": 2, "Amazon": 2, "Walmart": 2, "Brand Sites": 2,
    "Normalized Comparison": 10, "Opportunity Matrix": 3,
}

CSV_REQUIRED_COLUMNS = [
    "record_id", "channel", "product_name", "exclusion_reason",
    "payload_lb", "motor_w", "evidence_confidence", "opportunity_score",
]
CSV_REQUIRED_TERMS = [
    "ALI-EEC-BOX-60V", "BR-LECTRIC-XP-TRIKE2",
    "not_electric_cargo_tricycle", "moq_exceeds_pilot_limit",
]

NOTES_REQUIRED_TERMS = ["Alibaba.com", "Amazon", "Walmart", "Brand Sites", "未见"]
DECISION_REQUIRED_TERMS = ["BR-LECTRIC-XP-TRIKE2", "Alibaba", "未见"]
EXPECTED_CHANNELS_ORDER = ["Alibaba.com", "Amazon", "Walmart", "Brand Sites"]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def is_valid_image(p: Path) -> bool:
    if not p.exists():
        return False
    try:
        try:
            from PIL import Image  # type: ignore
            with Image.open(p) as im:
                im.verify()
            return True
        except ImportError:
            head = p.read_bytes()[:8]
            return head.startswith(b"\x89PNG\r\n\x1a\n") or head[:2] == b"\xff\xd8" or head[:4] == b"GIF8"
    except Exception:
        return False


def load_workbook(p: Path) -> tuple[Any, str]:
    if not p.exists():
        return None, f"missing {p.name}"
    try:
        import openpyxl  # type: ignore
        return openpyxl.load_workbook(str(p), data_only=True), ""
    except ImportError:
        return None, "openpyxl not installed"
    except Exception as e:
        return None, f"failed to load workbook: {e}"


def check_xlsx_parseable(p: Path) -> tuple[bool, str, Any]:
    wb, err = load_workbook(p)
    if wb is None:
        return False, err, None
    return True, "workbook loaded", wb


def check_xlsx_required_sheets(wb: Any) -> tuple[bool, str]:
    sheets = wb.sheetnames
    missing = [s for s in REQUIRED_SHEETS if s not in sheets]
    if missing:
        return False, f"missing sheets: {missing}"
    return True, f"all 6 sheets present: {sheets}"


def check_xlsx_min_rows(wb: Any) -> tuple[bool, str]:
    fails = []
    for sheet_name, min_rows in SHEET_MIN_ROWS.items():
        if sheet_name not in wb.sheetnames:
            fails.append(f"{sheet_name}: missing")
            continue
        ws = wb[sheet_name]
        n = ws.max_row or 0
        if n < min_rows:
            fails.append(f"{sheet_name}: {n} < {min_rows}")
    if fails:
        return False, "; ".join(fails)
    return True, "all sheets have min rows"


def check_notes_md(p: Path) -> tuple[bool, str]:
    if not p.exists():
        return False, "missing"
    text = p.read_text(encoding="utf-8-sig")
    if len(text) < 500:
        return False, f"length {len(text)} < 500"
    missing = [t for t in NOTES_REQUIRED_TERMS if t not in text]
    if missing:
        return False, f"missing terms: {missing}"
    return True, f"length={len(text)}, all terms present"


def check_evidence_manifest(p: Path) -> tuple[bool, str]:
    if not p.exists():
        return False, "missing file"
    try:
        data = load_json(p)
    except json.JSONDecodeError as e:
        return False, f"not parseable: {e}"
    channels = data.get("channels")
    if not isinstance(channels, list) or len(channels) != 4:
        return False, f"channels length {len(channels) if isinstance(channels, list) else 'N/A'}, need 4"
    for i, expected in enumerate(EXPECTED_CHANNELS_ORDER):
        ch = channels[i] if isinstance(channels[i], dict) else {}
        if ch.get("channel") != expected:
            return False, f"channels[{i}].channel={ch.get('channel')!r}, expected {expected!r}"
        if not isinstance(ch.get("sources"), list) or len(ch["sources"]) < 1:
            return False, f"channels[{i}].sources empty"
        if not isinstance(ch.get("records"), list) or len(ch["records"]) < 1:
            return False, f"channels[{i}].records empty"
        ss = ch.get("screenshot", "")
        if not isinstance(ss, str) or not ss.strip():
            return False, f"channels[{i}].screenshot empty"
    text = json.dumps(data, ensure_ascii=False)
    if "record_id" not in text:
        return False, "manifest text does not contain 'record_id'"
    return True, "manifest schema ok"


def check_csv_columns_and_terms(p: Path) -> tuple[bool, str]:
    if not p.exists():
        return False, "missing file"
    try:
        text = p.read_text(encoding="utf-8-sig")
        reader = csv.reader(StringIO(text))
        rows = list(reader)
    except Exception as e:
        return False, f"csv parse error: {e}"
    if not rows:
        return False, "empty csv"
    header = [h.strip() for h in rows[0]]
    missing_cols = [c for c in CSV_REQUIRED_COLUMNS if c not in header]
    if missing_cols:
        return False, f"missing columns: {missing_cols}"
    missing_terms = [t for t in CSV_REQUIRED_TERMS if t not in text]
    if missing_terms:
        return False, f"missing terms: {missing_terms}"
    return True, f"csv columns + terms ok ({len(rows)-1} data rows)"


def check_decision_matrix(p: Path) -> tuple[bool, str]:
    if not p.exists():
        return False, "missing file"
    try:
        data = load_json(p)
    except json.JSONDecodeError as e:
        return False, f"not parseable: {e}"
    if not isinstance(data, dict):
        return False, "not a dict"
    rec = data.get("recommended_record_id")
    if not isinstance(rec, str) or not rec.strip():
        return False, "recommended_record_id empty"
    for f in ("top_alternatives", "excluded_records", "channel_risks"):
        v = data.get(f)
        if not isinstance(v, list):
            return False, f"{f} not a list"
    text = json.dumps(data, ensure_ascii=False)
    missing = [t for t in DECISION_REQUIRED_TERMS if t not in text]
    if missing:
        return False, f"missing terms: {missing}"
    return True, "decision matrix ok"


def check_screenshots(evidence_dir: Path) -> tuple[bool, str]:
    missing = [n for n in SCREENSHOTS if not is_valid_image(evidence_dir / n)]
    if missing:
        return False, f"missing/invalid: {sorted(missing)}"
    return True, f"all {len(SCREENSHOTS)} screenshots valid"


def check_outputs_dir_clean(output_dir: Path) -> tuple[bool, str]:
    expected_root = {XLSX_NAME, NOTES_NAME, MANIFEST_NAME, CSV_NAME, DECISION_NAME, "evidence"}
    found_root = {p.name for p in output_dir.iterdir() if not p.name.startswith(".")}
    extras = found_root - expected_root
    if extras:
        return False, f"unexpected files/dirs in outputs/: {sorted(extras)}"
    evidence_dir = output_dir / "evidence"
    if evidence_dir.exists():
        found_evidence = {p.name for p in evidence_dir.iterdir() if not p.name.startswith(".")}
        extras_e = found_evidence - SCREENSHOTS
        if extras_e:
            return False, f"unexpected evidence/: {sorted(extras_e)}"
    return True, "outputs/ clean"


def run_llm_checks(
    llm_checks: list[dict],
    notes_path: Path,
    manifest_path: Path,
    csv_path: Path,
    decision_path: Path,
    expected_facts_path: Path | None = None,
) -> list[dict[str, Any]]:
    if not llm_checks:
        return []
    provider = os.environ.get("BENCH_LLM_JUDGE_PROVIDER")
    model = os.environ.get("BENCH_LLM_JUDGE_MODEL")
    if not provider or not model:
        return [{"id": c["id"], "passed": False, "reason": "BENCH_LLM_JUDGE_PROVIDER/MODEL not set",
                 "check_type": c.get("check_type", "llm_judge_boolean"), "llm_skipped": True} for c in llm_checks]

    try:
        from bench_core import llm_judge  # type: ignore
    except ImportError as exc:
        return [{"id": c["id"], "passed": False, "reason": f"llm_judge import failed: {exc}",
                 "check_type": c.get("check_type", "llm_judge_boolean"), "llm_skipped": True} for c in llm_checks]

    config = llm_judge.JudgeConfig(
        provider=provider, model=model,
        timeout_sec=int(os.environ.get("BENCH_LLM_JUDGE_TIMEOUT") or 240),
        base_url=os.environ.get("BENCH_LLM_JUDGE_BASE_URL"),
        api_key=os.environ.get("BENCH_LLM_JUDGE_API_KEY"),
    )

    notes = notes_path.read_text(encoding="utf-8-sig")[:20000] if notes_path.exists() else ""
    manifest = manifest_path.read_text(encoding="utf-8-sig")[:30000] if manifest_path.exists() else "{}"
    csv_text = csv_path.read_text(encoding="utf-8-sig")[:30000] if csv_path.exists() else ""
    decision = decision_path.read_text(encoding="utf-8-sig")[:15000] if decision_path.exists() else "{}"
    expected_facts_text = ""
    if expected_facts_path and expected_facts_path.exists():
        try:
            expected_facts_text = expected_facts_path.read_text(encoding="utf-8-sig")[:20000]
        except Exception:
            expected_facts_text = ""

    questions_block = "\n".join(f"  - id: {c['id']}\n    question: {c['description']}" for c in llm_checks)

    prompt = (
        "You are a strict benchmark verifier. Return JUDGE_SCHEMA `criteria: [{id, score, reason}]`. "
        "Each criterion is YES/NO: 1.0 if fully satisfied, 0.0 otherwise. No partial credit.\n\n"
        "Overall: `score` = sum/count; `passed` = (score >= 0.8); `summary` = one line.\n\n"
        f"Criteria:\n{questions_block}\n\n"
        "Reference facts (ground truth, agent did NOT see this):\n"
        f"{expected_facts_text or '(none)'}\n\n"
        f"=== {NOTES_NAME} ===\n{notes}\n\n"
        f"=== {MANIFEST_NAME} ===\n{manifest}\n\n"
        f"=== {CSV_NAME} ===\n{csv_text}\n\n"
        f"=== {DECISION_NAME} ===\n{decision}\n"
    )

    try:
        judgment = llm_judge.run_llm_judge(config, prompt)
    except Exception as e:
        return [{"id": c["id"], "passed": False, "reason": f"llm call failed: {e}",
                 "check_type": c.get("check_type", "llm_judge_boolean"), "llm_skipped": True} for c in llm_checks]

    criteria = judgment.get("criteria") or []
    by_id: dict[str, dict] = {}
    for item in criteria:
        if isinstance(item, dict) and isinstance(item.get("id"), str):
            by_id[item["id"]] = item

    results: list[dict[str, Any]] = []
    for c in llm_checks:
        item = by_id.get(c["id"])
        if not item:
            results.append({"id": c["id"], "passed": False, "reason": "LLM omitted",
                            "check_type": c.get("check_type", "llm_judge_boolean")})
            continue
        try:
            score = float(item.get("score", 0))
        except (TypeError, ValueError):
            score = 0.0
        results.append({"id": c["id"], "passed": score >= 0.5, "reason": str(item.get("reason", ""))[:300],
                        "check_type": c.get("check_type", "llm_judge_boolean"), "llm_score": score})
    return results


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--task-dir", required=True)
    ap.add_argument("--output-dir", default="")
    ap.add_argument("--reward-json", required=True)
    args = ap.parse_args()

    task_dir = Path(args.task_dir).resolve()
    reward_path = Path(args.reward_json)
    reward_path.parent.mkdir(parents=True, exist_ok=True)

    output_dir = Path(args.output_dir) if args.output_dir else task_dir / "outputs"
    evidence_dir = output_dir / "evidence"
    xlsx_path = output_dir / XLSX_NAME
    notes_path = output_dir / NOTES_NAME
    manifest_path = output_dir / MANIFEST_NAME
    csv_path = output_dir / CSV_NAME
    decision_path = output_dir / DECISION_NAME

    rubric = load_json(task_dir / "rubric.json")
    checks_spec = rubric.get("checks", [])

    det: list[dict[str, Any]] = []

    # --- Capability fold (capacity migration 2026-06-09) ---------------------
    # The 8 deterministic schema validations below all test ONE capability unit
    # ("produce every required deliverable in the mandated schema"): workbook
    # 6-sheet structure + min rows, notes shape, manifest fixed-order schema,
    # CSV columns+required-records, decision-JSON fields, screenshots present.
    # Per R1/R2 they collapse into a single `deliverables_schema_valid` check
    # (content correctness of each deliverable is graded independently by the 6
    # LLM capability checks). The negative restraint (`outputs_dir_*`, R6) stays
    # its own check. Every sub-validation is still run; the per-sub failure
    # detail is preserved in the `reason` string so diagnosis is unchanged.
    # new_check = AND(old atoms); binary pass is byte-for-byte unchanged.
    schema_atoms: list[tuple[bool, str]] = []

    ok, reason, wb = check_xlsx_parseable(xlsx_path)
    schema_atoms.append((ok, f"xlsx_exists_and_parseable: {reason}"))

    if wb is not None:
        ok, reason = check_xlsx_required_sheets(wb)
    else:
        ok, reason = False, "xlsx unavailable"
    schema_atoms.append((ok, f"xlsx_has_required_6_sheets: {reason}"))

    if wb is not None:
        ok, reason = check_xlsx_min_rows(wb)
    else:
        ok, reason = False, "xlsx unavailable"
    schema_atoms.append((ok, f"xlsx_sheet_min_row_counts: {reason}"))

    ok, reason = check_notes_md(notes_path)
    schema_atoms.append((ok, f"notes_md_exists_with_required_terms: {reason}"))

    ok, reason = check_evidence_manifest(manifest_path)
    schema_atoms.append((ok, f"evidence_manifest_parseable_and_4_channels_in_order: {reason}"))

    ok, reason = check_csv_columns_and_terms(csv_path)
    schema_atoms.append((ok, f"normalized_comparison_csv_has_required_columns_and_terms: {reason}"))

    ok, reason = check_decision_matrix(decision_path)
    schema_atoms.append((ok, f"decision_matrix_json_required_fields: {reason}"))

    ok, reason = check_screenshots(evidence_dir)
    schema_atoms.append((ok, f"all_4_evidence_screenshots_valid_images: {reason}"))

    schema_failures = [msg for sub_ok, msg in schema_atoms if not sub_ok]
    schema_ok = not schema_failures
    schema_reason = "all deliverables conform to schema" if schema_ok else "; ".join(schema_failures)
    det.append({"id": "deliverables_schema_valid", "passed": schema_ok, "reason": schema_reason, "check_type": "deterministic_exact"})

    if output_dir.exists():
        ok, reason = check_outputs_dir_clean(output_dir)
    else:
        ok, reason = False, "outputs/ dir missing"
    det.append({"id": "outputs_dir_has_only_declared_files", "passed": ok, "reason": reason, "check_type": "deterministic_exact"})

    llm_specs = [c for c in checks_spec if c.get("check_type") == "llm_judge_boolean"]
    private_dir = Path(os.environ.get("PRIVATE_DIR") or (task_dir / "private"))
    expected_facts_path = private_dir / "expected_facts.md"
    llm_results = run_llm_checks(llm_specs, notes_path, manifest_path, csv_path, decision_path, expected_facts_path)

    by_id = {r["id"]: r for r in det + llm_results}
    ordered = [by_id.get(c["id"], {"id": c["id"], "passed": False, "reason": "no result", "check_type": c.get("check_type", "")}) for c in checks_spec]

    passed = sum(1 for r in ordered if r["passed"])
    total = len(ordered)
    score = passed / total if total > 0 else 0.0
    n_llm_skipped = sum(1 for r in ordered if r.get("llm_skipped"))

    reward = {
        "schema_version": "2.0", "task_id": rubric.get("task_id"),
        "score": round(score, 4), "checks_passed": passed, "checks_total": total,
        "llm_checks_skipped": n_llm_skipped, "checks_breakdown": ordered,
        "reward": round(score, 4), "passed": passed == total, "source": "v2_checks_runner",
    }
    reward_path.write_text(json.dumps(reward, ensure_ascii=False, indent=2))
    print(json.dumps({k: reward[k] for k in ("score", "checks_passed", "checks_total", "llm_checks_skipped")}))

    return 0


if __name__ == "__main__":
    sys.exit(main())
